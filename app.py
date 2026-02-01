import streamlit as st
import google.generativeai as genai
import pandas as pd
import holidays
import re
import calendar
from datetime import date, datetime
from PIL import Image, ImageOps
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.chart import BarChart, Reference
from io import BytesIO

# --- 1. KONFIGURACJA AI ---
@st.cache_resource
def get_working_model():
    try:
        genai.configure(api_key=st.secrets["GOOGLE_API_KEY"])
        available = [m.name for m in genai.list_models() if 'generateContent' in m.supported_generation_methods]
        # Wybieramy stabilny model 1.5 Flash dla najlepszych limitów
        for target in ["models/gemini-1.5-flash-8b", "models/gemini-1.5-flash"]:
            if target in available: return genai.GenerativeModel(target), target
        return None, "Brak połączenia"
    except Exception: return None, "Błąd"

model, active_model_name = get_working_model()

# --- 2. FUNKCJE POMOCNICZE ---
M_LIST = ["Styczeń", "Luty", "Marzec", "Kwiecień", "Maj", "Czerwiec", "Lipiec", "Sierpień", "Wrzesień", "Październik", "Listopad", "Grudzień"]

def get_working_info(year, month):
    pl_hols = holidays.Poland(years=year)
    num_days = calendar.monthrange(year, month)[1]
    working_days = 0
    holiday_list = []
    for day in range(1, num_days + 1):
        curr_date = date(year, month, day)
        if curr_date in pl_hols and curr_date.weekday() < 5:
            holiday_list.append(f"{day} {M_LIST[month-1]} - {pl_hols.get(curr_date)}")
        if curr_date.weekday() < 5 and curr_date not in pl_hols:
            working_days += 1
    return working_days * 8, holiday_list

def get_day_name(year, month, day):
    dni = ["Pon", "Wto", "Śro", "Czw", "Pią", "Sob", "Nie"]
    try: return dni[date(year, month, day).weekday()]
    except: return ""

def process_excel_save(uploaded_file, new_data, pil_image):
    """Główna funkcja zapisu: zachowuje stare zdjęcia i aktualizuje dane."""
    if uploaded_file:
        # Ładujemy istniejący plik
        wb = openpyxl.load_workbook(BytesIO(uploaded_file.read()))
    else:
        # Tworzymy nowy jeśli nic nie wgrano
        wb = openpyxl.Workbook()
        wb.active.title = "Zarobki"
        ws = wb.active
        ws.append(["Rok", "Miesiąc", "Godziny Suma", "Norma", "Nadgodziny", "Stawka", "Dni Urlopu", "Suma PLN", "Grafik"])

    ws = wb["Zarobki"]
    
    # Szukamy czy ten rok i miesiąc już są w tabeli
    found_row = None
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=1).value) == str(new_data["Rok"]) and \
           str(ws.cell(row=row, column=2).value) == str(new_data["Miesiąc"]):
            found_row = row
            break
    
    target_row = found_row if found_row else ws.max_row + 1
    
    # Wpisujemy dane tekstowe
    ws.cell(row=target_row, column=1, value=new_data["Rok"])
    ws.cell(row=target_row, column=2, value=new_data["Miesiąc"])
    ws.cell(row=target_row, column=3, value=new_data["Godziny Suma"])
    ws.cell(row=target_row, column=4, value=new_data["Norma"])
    ws.cell(row=target_row, column=5, value=new_data["Nadgodziny"])
    ws.cell(row=target_row, column=6, value=new_data["Stawka"])
    ws.cell(row=target_row, column=7, value=new_data["Dni Urlopu"])
    ws.cell(row=target_row, column=8, value=new_data["Suma PLN"])

    # Dodajemy zdjęcie
    if pil_image:
        img_temp = pil_image.copy()
        img_temp.thumbnail((1600, 2000))
        img_buf = BytesIO()
        img_temp.save(img_buf, format="PNG")
        img_xl = OpenpyxlImage(img_buf)
        img_xl.width, img_xl.height = 80, 105
        ws.add_image(img_xl, f'I{target_row}')
        ws.row_dimensions[target_row].height = 80
        ws.column_dimensions['I'].width = 15

    # Aktualizacja Statystyk i Wykresów
    if "Statystyki" in wb.sheetnames: del wb["Statystyki"]
    ws_stats = wb.create_sheet("Statystyki")
    
    # Pobieramy dane do wykresów z arkusza Zarobki
    data_for_stats = []
    for r in range(1, ws.max_row + 1):
        data_for_stats.append([ws.cell(r, c).value for c in range(1, 9)])
    
    df_temp = pd.DataFrame(data_for_stats[1:], columns=data_for_stats[0])
    df_temp['M_Idx'] = df_temp['Miesiąc'].apply(lambda x: M_LIST.index(x) if x in M_LIST else 0)
    df_temp = df_temp.sort_values(['Rok', 'M_Idx'])
    
    # Zapisujemy tabelę statystyk
    for r_idx, row in enumerate(df_temp[['Miesiąc', 'Godziny Suma', 'Nadgodziny', 'Suma PLN']].values, 1):
        if r_idx == 1: ws_stats.append(['Miesiąc', 'Godziny Suma', 'Nadgodziny', 'Suma PLN'])
        ws_stats.append(list(row))

    # Wykresy
    max_s = ws_stats.max_row
    if max_s > 1:
        for i, title in enumerate(["Suma Godzin", "Nadgodziny", "Zarobki PLN"], 2):
            c = BarChart()
            c.title = title
            c.add_data(Reference(ws_stats, min_col=i, min_row=1, max_row=max_s), titles_from_data=True)
            c.set_categories(Reference(ws_stats, min_col=1, min_row=2, max_row=max_s))
            ws_stats.add_chart(c, f"F{2 + (i-2)*15}")

    final_out = BytesIO()
    wb.save(final_out)
    return final_out.getvalue()

# --- 3. INTERFEJS ---
st.set_page_config(page_title="Kalkulator czasu pracy", layout="wide")

with st.sidebar:
    st.title("📂 Baza danych")
    uploaded_file = st.file_uploader("Wgraj swój plik Excel (.xlsx):", type="xlsx")
    st.divider()
    st.header("⚙️ Ustawienia")
    cur_yr = datetime.now().year
    rok = st.selectbox("Rok:", list(range(2024, cur_yr + 11)), index=list(range(2024, cur_yr + 11)).index(cur_yr))
    m_nazwa = st.selectbox("Miesiąc:", M_LIST, index=datetime.now().month-1)
    stawka = st.number_input("Stawka podstawowa (zł/h):", value=25.0)
    dodatek = st.number_input("Dodatek za nadgodziny (zł):", value=15.0)

st.title("Kalkulator czasu pracy")
st.caption(f"🤖 Aktywny model: `{active_model_name}`")

m_idx = M_LIST.index(m_nazwa) + 1
norma_h, swieta = get_working_info(rok, m_idx)

with st.expander(f"📅 Norma dla {m_nazwa} {rok}", expanded=False):
    st.write(f"Wymiar: **{norma_h} h**")
    for s in swieta: st.write(f"• {s}")

plik = st.file_uploader("Wgraj zdjęcie grafiku:", type=['jpg', 'jpeg', 'png'])

if plik:
    img = ImageOps.exif_transpose(Image.open(plik))
    st.image(img, width=300)
    
    if st.button("🚀 SKANUJ GRAFIK"):
        with st.spinner("AI analizuje grafik..."):
            try:
                prompt = """Odczytaj kolumnę 'Ilość godzin' dla dni 1-31. Format: 1: [wart], 2: [wart]... 
                ZASADY: 'x', 'X', '-' to 0. 'URL', 'Urlop' to U."""
                response = model.generate_content([prompt, img])
                pairs = re.findall(r"(\d+):\s*([0-9.UuXx-]+)", response.text)
                d_list = [0.0]*31
                u_list = []
                for d, v in pairs:
                    dn = int(d)
                    if dn <= 31:
                        val = v.upper().strip()
                        if 'U' in val:
                            d_list[dn-1] = 8.0
                            u_list.append(dn)
                        elif val in ['X', '-', '']: d_list[dn-1] = 0.0
                        else:
                            try: d_list[dn-1] = float(re.findall(r"(\d+(?:\.\d+)?)", val)[0])
                            except: pass
                st.session_state['dni_lista'] = d_list
                st.session_state['url_dni'] = u_list
                st.session_state['last_img'] = img
                st.success("✅ Odczytano!")
            except Exception as e: st.error(f"Błąd AI: {e}")

if 'dni_lista' in st.session_state:
    st.divider()
    num_d = calendar.monthrange(rok, m_idx)[1]
    st.subheader("📝 Korekta i wyniki")
    
    sel_url = st.multiselect("Dni urlopowe (8h):", range(1, num_d + 1), 
                             default=st.session_state.get('url_dni', []),
                             format_func=lambda x: f"Dzień {x} ({get_day_name(rok, m_idx, x)})")

    popr = []
    c_l, c_r = st.columns(2)
    for i in range(num_d):
        dn = i + 1
        with (c_l if i < num_d/2 else c_r):
            d_init = 8.0 if dn in sel_url else st.session_state['dni_lista'][i]
            v = st.number_input(f"Dz {dn} ({get_day_name(rok, m_idx, dn)})", value=float(d_init), step=0.5, key=f"k_{i}")
            popr.append(v)
    
    suma_h = sum(popr)
    nadgodziny = max(0.0, suma_h - norma_h)
    total = (suma_h * stawka) + (nadgodziny * dodatek)
    
    # --- METRYKI NA STRONIE GŁÓWNEJ ---
    st.divider()
    col1, col2, col3 = st.columns(3)
    col1.metric("Suma godzin", f"{suma_h} h")
    col2.metric("Nadgodziny", f"{nadgodziny} h")
    col3.metric("Wypłata BRUTTO", f"{total:,.2f} zł")
    st.divider()

    if st.button("📊 Zapisz i przygotuj plik Excel"):
        excel_ready = process_excel_save(uploaded_file, {
            "Rok": rok, "Miesiąc": m_nazwa, "Godziny Suma": suma_h,
            "Norma": norma_h, "Nadgodziny": nadgodziny,
            "Stawka": stawka, "Dni Urlopu": len(sel_url), "Suma PLN": round(total, 2)
        }, st.session_state.get('last_img'))
        
        st.session_state['excel_ready'] = excel_ready
        st.success("✅ Plik zaktualizowany (stare zdjęcia zachowane)!")

    if 'excel_ready' in st.session_state:
        st.download_button("📥 POBIERZ AKTUALNY EXCEL", data=st.session_state['excel_ready'], 
                           file_name=f"zarobki_{rok}_{m_nazwa}.xlsx")
